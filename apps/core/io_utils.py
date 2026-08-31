"""
Shared import/export utilities used across all apps.
"""
import csv
import io
import re
from decimal import Decimal, InvalidOperation
from datetime import datetime

from django.http import HttpResponse


# ── Response helpers ────────────────────────────────────────────────────────

def csv_response(filename: str) -> HttpResponse:
    """Return an HttpResponse ready for CSV download with UTF-8 BOM (Excel-safe)."""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write('﻿')  # single BOM — charset=utf-8 never prepends its own
    return response


def csv_writer(response) -> csv.writer:
    return csv.writer(response)


# ── Parsing helpers ─────────────────────────────────────────────────────────

def parse_uploaded_file(file) -> tuple[list[dict], str | None]:
    """
    Parse an uploaded CSV or Excel file.
    Returns (rows, error_message).
    rows is a list of dicts keyed by the header row values.
    """
    name = file.name.lower()
    if not name.endswith(('.csv', '.xlsx', '.xls')):
        return [], 'نوع الملف غير مدعوم. الرجاء رفع ملف Excel أو CSV'

    try:
        if name.endswith('.csv'):
            decoded = file.read().decode('utf-8-sig')
            decoded = decoded.lstrip('﻿')  # strip any extra BOMs left after decode
            reader = csv.DictReader(io.StringIO(decoded))
            return list(reader), None
        else:
            import openpyxl
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            headers = [
                str(c.value).strip().lstrip('﻿') if c.value is not None else ''
                for c in ws[1]
            ]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(v is None for v in row):
                    continue
                rows.append(dict(zip(headers, row)))
            return rows, None
    except Exception as exc:
        return [], f'تعذّر قراءة الملف: {exc}'


# ── Arabic numeral normalisation ────────────────────────────────────────────

_ARABIC_INDIC = str.maketrans('٠١٢٣٤٥٦٧٨٩', '0123456789')


def _normalize_num(text: str) -> str:
    """Convert Arabic-Indic digits and remove thousands separators."""
    return text.translate(_ARABIC_INDIC).replace(',', '').replace('،', '').strip()


# ── Field extraction ────────────────────────────────────────────────────────

_EMPTY = {'', 'none', 'nan', 'null', 'n/a', '-', '—'}


def _clean_str(val) -> str:
    if val is None:
        return ''
    s = str(val).strip().lstrip('﻿')
    return '' if s.lower() in _EMPTY else s


def get(row: dict, *keys, default='') -> str:
    """Return the first non-empty value matching any of the given keys (Arabic or English)."""
    for key in keys:
        s = _clean_str(row.get(key))
        if s:
            return s
    return default


def smart_get(row: dict, field: str, mapping: dict, *fallback_keys, default='') -> str:
    """
    Extract a field value using AI-generated header mapping first,
    then fall back to direct key lookup via fallback_keys.

    mapping: {actual_file_header: canonical_field_name}  (from smart_map_headers)
    """
    for actual_header, canonical in mapping.items():
        if canonical == field:
            s = _clean_str(row.get(actual_header))
            if s:
                return s
    return get(row, *fallback_keys, default=default)


# ── Type coercions ──────────────────────────────────────────────────────────

def safe_decimal(val, default=Decimal('0')) -> Decimal:
    try:
        return Decimal(_normalize_num(str(val)))
    except (InvalidOperation, ValueError, TypeError):
        return default


def safe_date(val, default=None):
    if val is None:
        return default
    if isinstance(val, datetime):
        return val.date()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(str(val).strip(), fmt).date()
        except ValueError:
            pass
    return default


def bool_from_str(val) -> bool:
    return _clean_str(val).lower() in ('نعم', 'yes', '1', 'true', 'y', '✓', 'صح', 'صحيح', 'active', 'نشط')


def clean_phone(val) -> str | None:
    """Strip formatting from a phone number; return None if empty."""
    s = re.sub(r'[\s\-().+]', '', _clean_str(val))
    s = _normalize_num(s)
    return s or None


def clean_email(val) -> str | None:
    """Lowercase and strip an email address; return None if empty."""
    s = _clean_str(val).lower()
    return s if '@' in s else None
